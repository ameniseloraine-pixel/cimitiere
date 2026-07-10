"""
API Reporting — Tableaux de bord, statistiques, exports CSV/Excel
"""

import csv
import io
from datetime import date
from typing import Optional

from django.http import HttpResponse
from ninja import Router
from apps.users.api import auth

router = Router()


@router.get("/dashboard", auth=auth)
def dashboard(request):
    """
    Données complètes pour le tableau de bord admin.
    - Taux d'occupation global et par bloc
    - Revenus financiers
    - Alertes actives
    """
    from apps.cartographie.models import Caveau, StatutCaveau
    from apps.terrain.models import Bloc
    from apps.finance.models import Facture, StatutFacture
    from apps.concessions.models import Concession
    from django.db.models import Count, Sum

    if not request.auth.peut_voir_finances and not request.auth.peut_modifier_carte:
        return {"message": "Accès limité"}

    # Occupation globale
    total = Caveau.objects.count()
    occupes = Caveau.objects.filter(statut=StatutCaveau.OCCUPE).count()
    reserves = Caveau.objects.filter(statut=StatutCaveau.RESERVE).count()
    disponibles = Caveau.objects.filter(statut=StatutCaveau.DISPONIBLE).count()

    # Occupation par bloc
    blocs_stats = []
    for bloc in Bloc.objects.prefetch_related("caveaux").all():
        total_bloc = bloc.caveaux.count()
        occupes_bloc = bloc.caveaux.filter(statut=StatutCaveau.OCCUPE).count()
        taux = round(occupes_bloc / total_bloc * 100, 1) if total_bloc else 0
        blocs_stats.append({
            "bloc": f"{bloc.zone.code}/{bloc.code}",
            "total": total_bloc,
            "occupes": occupes_bloc,
            "taux_occupation_pct": taux,
            "alerte_saturation": taux >= 90,
        })

    # Finances (si permission)
    finances = {}
    if request.auth.peut_voir_finances:
        from django.utils import timezone
        ce_mois = timezone.now().replace(day=1)
        finances = {
            "revenus_total": float(
                Facture.objects.filter(statut=StatutFacture.PAYEE)
                .aggregate(Sum("montant_total"))["montant_total__sum"] or 0
            ),
            "revenus_ce_mois": float(
                Facture.objects.filter(
                    statut=StatutFacture.PAYEE,
                    date_emission__gte=ce_mois
                ).aggregate(Sum("montant_total"))["montant_total__sum"] or 0
            ),
            "factures_en_retard": Facture.objects.filter(
                statut=StatutFacture.EN_RETARD
            ).count(),
        }

    # Alertes concessions
    alertes_expiration = sum(
        1 for c in Concession.objects.filter(statut="ACTIVE")
        if c.necessite_alerte
    )

    return {
        "occupation": {
            "total": total,
            "occupes": occupes,
            "reserves": reserves,
            "disponibles": disponibles,
            "taux_occupation_pct": round(occupes / total * 100, 1) if total else 0,
            "jauge_saturation": round((occupes + reserves) / total * 100, 1) if total else 0,
        },
        "par_bloc": blocs_stats,
        "finances": finances,
        "alertes": {
            "concessions_expirant_bientot": alertes_expiration,
        },
    }


@router.get("/export/registre-funeraire", auth=auth)
def export_registre_csv(request, format: str = "csv"):
    """
    Export du registre funéraire en CSV ou Excel.
    Réservé aux rôles Admin et Secrétariat.
    """
    if not request.auth.peut_voir_finances:
        return HttpResponse("Permission insuffisante", status=403)

    from apps.reservations.models import Reservation, StatutReservation

    reservations = Reservation.objects.filter(
        statut=StatutReservation.VALIDEE
    ).select_related("defunt", "caveau__bloc__zone", "client").order_by("-date_validation")

    if format == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registre Funéraire"

        # En-têtes
        headers = [
            "N° Dossier", "Défunt (Nom)", "Défunt (Prénom)",
            "Date Décès", "Date Inhumation", "Caveau", "Zone",
            "Client", "Email Client", "Date Validation",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="2d6a4f")

        for row, res in enumerate(reservations, 2):
            ws.append([
                res.numero_dossier,
                res.defunt.nom,
                res.defunt.prenom,
                str(res.defunt.date_deces),
                str(res.date_inhumation_souhaitee or ""),
                res.caveau.reference_complete,
                res.caveau.bloc.zone.code,
                res.client.nom_complet,
                res.client.email,
                str(res.date_validation.date() if res.date_validation else ""),
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="registre_funeraire.xlsx"'
        return response

    # CSV par défaut
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="registre_funeraire.csv"'
    response.write("\ufeff")  # BOM UTF-8 pour Excel

    writer = csv.writer(response, delimiter=";")
    writer.writerow([
        "N° Dossier", "Défunt Nom", "Défunt Prénom",
        "Date Décès", "Date Inhumation", "Référence Caveau", "Zone",
        "Client", "Email", "Date Validation",
    ])
    for res in reservations:
        writer.writerow([
            res.numero_dossier, res.defunt.nom, res.defunt.prenom,
            res.defunt.date_deces, res.date_inhumation_souhaitee or "",
            res.caveau.reference_complete, res.caveau.bloc.zone.code,
            res.client.nom_complet, res.client.email,
            res.date_validation.date() if res.date_validation else "",
        ])
    return response
